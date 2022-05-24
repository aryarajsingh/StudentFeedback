from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from sklearn.feature_extraction.text import CountVectorizer
from nltk.tokenize import RegexpTokenizer
from sklearn.model_selection import train_test_split
from sklearn.naive_bayes import MultinomialNB
from sklearn.metrics import classification_report, confusion_matrix
import pandas as pd

from textblob import TextBlob

import seaborn as sn
from wordcloud import WordCloud


import matplotlib.pyplot as plt
from io import StringIO
import numpy as np
import pickle

from openpyxl import Workbook
import openpyxl


def index(request):
    if request.method == 'POST':

        if 'shorts' in request.POST:
            return redirect('updateshorts')

        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request,username=username,password=password)

        if user is not None:
            login(request,user)
            return redirect('home')
        else:
            print("wrong credentials.")

    return render(request=request,
                  template_name='index.html')

def home(request):

    if request.method == 'POST':

        path = request.POST.get('path')
        data = pd.read_excel(path)

        enum_list = list(zip(data['feedback'], data['points']))
        pickle_out = open("dict.pickle", "wb")
        pickledata = {'enum_list':enum_list,
                      'path':path}
        pickle.dump(pickledata, pickle_out)
        pickle_out.close()

        return redirect('ShowTableData')
    return render(request=request,
                  template_name='home.html')

def dashboard(request):
    return render(request=request,
                  template_name='dashboard.html')

def ShowTableData(request):
    if request.method == 'POST':

        pickle_in = open("dict.pickle", "rb")
        pickledata = pickle.load(pickle_in)
        path = pickledata.get('path')
        data = pd.read_excel(path)

        token = RegexpTokenizer(r'[a-zA-Z0-9]+')
        cv = CountVectorizer(lowercase=True, stop_words='english', ngram_range=(1, 1), tokenizer=token.tokenize)
        text_counts = cv.fit_transform(data['feedback'])

        X_train, X_test, y_train, y_test = train_test_split(text_counts, data['points'], test_size=0.3, random_state=1)

        clf = MultinomialNB().fit(X_train, y_train)

        predicted = clf.predict(X_test)

        confusionMatrix = confusion_matrix(y_test, predicted)
        classificationReport = classification_report(y_test, predicted,output_dict=True)

        negativeReview = classificationReport['-1']
        neutralReview = classificationReport['0']
        positiveReview = classificationReport['1']

        accuracy = classificationReport['accuracy']
        macro_avg = classificationReport['macro avg']
        weighted_avg = classificationReport['weighted avg']

        def return_graph(matrix):

            df_cm = pd.DataFrame(matrix, index=[i for i in "ABC"],
                                 columns=[i for i in "ABC"])

            fig = plt.figure(figsize=(5, 5))

            sn.heatmap(df_cm, annot=True)
            plt.xlabel("Predicted")
            plt.ylabel('Actual')

            imgdata = StringIO()
            fig.savefig(imgdata, format='svg')
            imgdata.seek(0)

            return imgdata.getvalue()

        def show_wordcloud(data, title=None):
            wordcloud = WordCloud(
                background_color='white',
                max_words=1000,
                max_font_size=72,
                scale=5,
                random_state=42
            ).generate(str(data))



            clound_fig = plt.figure(figsize=(5, 5))
            plt.axis('off')
            plt.imshow(wordcloud)

            imagedata2 = StringIO()
            clound_fig.savefig(imagedata2, format='svg')
            imagedata2.seek(0)

            return imagedata2.getvalue()

        def show_piechart(data):
            positive_review = 0
            negative_review = 0
            neutral_review = 0

            for i in data['points']:
                if i == 1:
                    positive_review += 1

                if i == -1:
                    negative_review += 1
                if i == 0:
                    neutral_review += 1

            y = np.array([positive_review, negative_review, neutral_review])
            mylabels = ["Positive Reviews", "Negative Reviews", "Neutral Reviews"]
            myexplode = [0.1, 0, 0]

            pie_fig = plt.figure(figsize=(5, 5))
            plt.pie(y, labels=mylabels, autopct='%1.1f%%', explode=myexplode, shadow=True)

            imagedata3 = StringIO()
            pie_fig.savefig(imagedata3, format='svg')
            imagedata3.seek(0)

            return imagedata3.getvalue()

        context = {'graph': return_graph(confusionMatrix),
                   'wordcloud': show_wordcloud(data['feedback']),
                   'piechart': show_piechart(data),
                   'classificationReport':classificationReport,
                   'negativeReviewPrecision':negativeReview["precision"],
                   'negativeReviewRecall':negativeReview["recall"],
                   'negativeReviewF1Score':negativeReview["f1-score"],
                   'negativeReviewSupport':negativeReview["support"],

                   'neutralReviewPrecision': neutralReview["precision"],
                   'neutralReviewRecall': "{:.2f}".format(neutralReview["recall"]),
                   'neutralReviewF1Score': neutralReview["f1-score"],
                   'neutralReviewSupport': neutralReview["support"],

                   'positiveReviewPrecision': "{:.2f}".format(positiveReview["precision"]),
                   'positiveReviewRecall': positiveReview["recall"],
                   'positiveReviewF1Score': "{:.2f}".format(positiveReview["f1-score"]),
                   'positiveReviewSupport': positiveReview["support"],
                   'accuracy':"{:.2f}".format(accuracy),

                   'macro_avg_precision':"{:.2f}".format(macro_avg["precision"]),
                   'macro_avg_recall': "{:.2f}".format(macro_avg["recall"]),
                   'macro_avg_F1Score': "{:.2f}".format(macro_avg["f1-score"]),
                   'macro_avg_Support': macro_avg["support"],

                   'weighted_avg_precision':"{:.2f}".format(weighted_avg["precision"]),
                   'weighted_avg_recall': "{:.2f}".format(weighted_avg["recall"]),
                   'weighted_avg_F1Score': "{:.2f}".format(weighted_avg["f1-score"]),
                   'weighted_avg_Support': weighted_avg["support"]}

        return render(request=request,
                      template_name='dashboard.html',
                      context=context)

    pickle_in = open("dict.pickle", "rb")
    pickledata = pickle.load(pickle_in)
    enum_list = pickledata.get('enum_list')
    print(pickledata)
    return render(request=request,
                  template_name='showtabledata.html',
                  context={'enum_list':enum_list,
                           'enum_list_len':len(enum_list)})

def studentfeedback(request):
    if request.method == 'POST':
        feedback_data = request.POST.get('studentfeedback')
        data = TextBlob(feedback_data)
        points = data.sentiment.polarity




        if points > 0.5:
            print("Positive ",1)
            final_points=1
        elif points<0.5 and points>0:
            print("Average ",0)
            final_points=0
        else:
            print("Negative ",-1)
            final_points=-1

        print(feedback_data)

        #

        path = "studentfeedback.xlsx"
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active

        max_row = sheet_obj.max_row

        data = []
        for i in range(1, max_row + 1):
            cell_obj1 = sheet_obj.cell(row=i, column=1)
            cell_obj2 = sheet_obj.cell(row=i, column=2)

            data.append([cell_obj1.value, cell_obj2.value])

        print(data)

        book = Workbook()
        sheet = book.active

        data.append([final_points, str(feedback_data)])

        for row in data:
            sheet.append(row)

        book.save('studentfeedback.xlsx')



        return redirect('index')
    return render(request=request,
                  template_name='studentfeedback.html')